from datetime import datetime
from elasticsearch import Elasticsearch
from openpyxl import load_workbook
import json

wb = load_workbook(filename = 'titanic3.xlsx')
ws = wb['titanic3']
to_json = []
r_index = 0

keys = [x.internal_value for x in ws[1]]
for row in range(ws.max_row):
    if row == 0:
        continue
    new_row = {}
    new_row['index'] = r_index
    for index, cell in enumerate(ws[row + 1]):
        new_row[keys[index]] = cell.internal_value
    to_json.append(new_row)
    r_index += 1

es = Elasticsearch() #defaults to localhost:9200
es.indices.delete(index='titanic', ignore=[400, 404])
es.indices.create(index='titanic', ignore=400)
es.indices.put_mapping(
    index="titanic",
    doc_type="people",
    body={
        "properties" : {
          "age" : {"type" : "float"},
          "boat" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "body" : {"type" : "long"},
          "cabin" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "embarked" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "fare" : {"type" : "float"},
          "home" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "index" : {"type" : "long"},
          "name" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "parch" : {"type" : "long"},
          "pclass" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "sex" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    },
          "sibsp" : {"type" : "long"},
          "survived" : {"type" : "long"},
          "ticket" : {"type" : "text",
                    "fields": {
                        "keyword": {
                          "type":  "keyword",
                          "ignore_above" : 256
                                    }
                               }
                    }
        }
    }
)

# For a small dataset, you can create each document individually
# Otherwise use es.bulk 

for item in to_json:
    es.index(index='titanic', doc_type='people', body=json.dumps(item))
